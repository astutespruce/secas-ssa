from openpyxl.styles import (
    Alignment,
    Font,
    NamedStyle,
    Border,
    Side,
    PatternFill,
)
from openpyxl.utils.cell import get_column_letter

# Guess at how many characters fit into a column width measurement
CHAR_PER_WIDTH_UNIT = 1.7

### Create named styles for formatting cells
font_bold = Font(bold=True)


alignment_left_wrap = Alignment(horizontal="left", wrap_text=True)
alignment_center_wrap = Alignment(horizontal="center", wrap_text=True)

default_header_border = Border(
    bottom=Side(border_style="medium", color="000000"),
)
default_cell_border = Border(
    bottom=Side(border_style="thin", color="AAAAAA"),
    left=Side(border_style="thin", color="DDDDDD"),
    right=Side(border_style="thin", color="DDDDDD"),
)

# Note: all cells are setup to wrap text
left_header_style = NamedStyle(
    name="Left Header Style",
    font=font_bold,
    alignment=alignment_left_wrap,
    border=default_header_border,
)

center_header_style = NamedStyle(
    name="Center Header Style",
    font=font_bold,
    alignment=alignment_center_wrap,
    border=default_header_border,
)

value_style = NamedStyle(
    name="Value Style",
    alignment=alignment_left_wrap,
    border=default_cell_border,
)

even_row_bg = PatternFill("solid", fgColor="00F6F6F6")

analysis_unit_divider = Border(
    bottom=Side(border_style="medium", color="AAAAAA"),
    left=Side(border_style="thin", color="DDDDDD"),
    right=Side(border_style="thin", color="DDDDDD"),
)

description_font = Font(color="999999")


def set_cell_styles(ws, breaks=None, area_columns=None, percent_columns=None):
    area_columns = area_columns or []
    percent_columns = percent_columns or []

    for col_idx, col in enumerate(ws.columns):
        col[0].style = center_header_style

        for i, cell in enumerate(col[1:]):
            cell.style = value_style
            value = cell.value
            is_int = isinstance(value, (float, int)) and int(value) == value

            if col_idx in area_columns:
                if is_int:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"
            elif col_idx in percent_columns:
                if is_int:
                    cell.number_format = "0%"
                else:
                    cell.number_format = "0.00%"

            if i % 2 == 1:
                cell.fill = even_row_bg

    ws["A1"].style = left_header_style

    if breaks is not None:
        # add a stronger line between analysis units
        for col in ws.columns:
            for line in breaks:
                col[line].border = analysis_unit_divider


def set_column_widths(ws, widths):
    for i, width in enumerate(widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = width
