import pandas as pd

from openpyxl.styles import Font, Color

from analysis.constants import DATASETS

from api.report.style import (
    description_font,
    alignment_left_wrap,
    set_cell_styles,
    set_column_widths,
)


def add_data_note(ws, content, columns=None):
    columns = columns or len(list(ws.columns)) - 1
    # add data note below data table
    row_index = len(list(ws.rows)) + 3
    ws.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index + 1,
        end_column=columns + 1,
    )
    cell_index = f"A{row_index}"
    ws[cell_index].value = content
    ws[cell_index].font = description_font
    ws[cell_index].alignment = alignment_left_wrap


def add_data_details_sheet(xlsx, datasets):
    # Keep original order so it matches sheets
    pd.DataFrame([DATASETS[dataset] for dataset in DATASETS if dataset in datasets])[
        [
            "name",
            "sheet_name",
            "source",
            "date",
            "description",
            "citation",
            "url",
        ]
    ].rename(
        columns={
            "name": "Name",
            "sheet_name": "Sheet",
            "source": "Data source",
            "date": "Date",
            "description": "Description",
            "citation": "Citation",
            "url": "URL",
        }
    ).to_excel(
        xlsx, sheet_name="Data details", index=False
    )
    ws = xlsx.sheets["Data details"]
    set_column_widths(ws, [16, 16, 18, 8, 48, 32, 40])
    set_cell_styles(ws)
    for cell in list(ws.columns)[-1][1:]:
        cell.hyperlink = cell.value
        cell.font = Font(color=Color(index=4))
